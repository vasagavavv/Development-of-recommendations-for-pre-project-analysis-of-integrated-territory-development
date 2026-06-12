import osmnx as ox
import geopandas as gpd
from shapely.geometry import Point
from typing import Optional, Dict, Any
import os
import csv
from datetime import datetime
from openpyxl import load_workbook


def only_polygons(gdf: Optional[gpd.GeoDataFrame]) -> Optional[gpd.GeoDataFrame]:
    """Return GeoDataFrame with only Polygon/MultiPolygon geometries or None/empty as-is."""
    if gdf is None or gdf.empty or "geometry" not in gdf.columns:
        return gdf
    g = gdf[~gdf.geometry.isna()].copy()
    if g.empty:
        return g
    return g[g.geometry.geom_type.isin(["Polygon", "MultiPolygon"]) ]


def point_in_polygons(gdf: Optional[gpd.GeoDataFrame], point: Point) -> bool:
    """Check whether point is covered by any polygon in gdf.

    Uses unary_union for faster single-point checks when possible.
    """
    g = only_polygons(gdf)
    if g is None or g.empty:
        return False
    try:
        union = g.geometry.unary_union
        return bool(union.covers(point))
    except Exception:
        # fallback to elementwise check
        return bool(g.geometry.apply(lambda geom: geom.covers(point)).any())


def get_features(lat: float, lon: float, tags: Dict[str, str], dist: int):
    """Safe wrapper around osmnx.features_from_point; returns GeoDataFrame or None."""
    try:
        gdf = ox.features_from_point((lat, lon), tags, dist=dist)
        return gdf
    except Exception as e:
        # don't crash the caller; return None and let caller decide
        print(f"Ошибка получения данных OSM для тегов {tags}: {e}")
        return None


def classify_point(proj_lat: float, proj_lon: float, R_CAT: int = 1500, raise_on_fail: bool = True) -> Dict[str, Any]:
    """Classify point by OSM polygons.

    Returns a dict with keys: in_park_poly, in_wood_poly, category (or None), details.
    If raise_on_fail is True and no category matched, raises ValueError.
    """
    pt = Point(proj_lon, proj_lat)

    tags_park = {"leisure": "park"}
    tags_wood = {"natural": "wood"}

    gdf_park = get_features(proj_lat, proj_lon, tags_park, R_CAT)
    gdf_wood = get_features(proj_lat, proj_lon, tags_wood, R_CAT)

    # Ensure CRS is WGS84 (EPSG:4326) for reliable geometric tests
    for gdf in (gdf_park, gdf_wood):
        if gdf is None or gdf.empty:
            continue
        try:
            if gdf.crs is None:
                gdf.set_crs(epsg=4326, inplace=True)
            else:
                gdf.to_crs(epsg=4326, inplace=True)
        except Exception:
            # If CRS operations fail, continue — point checks may still work if geometries are lat/lon
            pass

    in_park_poly = point_in_polygons(gdf_park, pt)
    in_wood_poly = point_in_polygons(gdf_wood, pt)

    details = {
        "in_park_poly": in_park_poly,
        "in_wood_poly": in_wood_poly,
        "R_CAT": R_CAT,
    }

    if in_park_poly or in_wood_poly:
        category = "park"
        details["category"] = category
        return details

    # no match
    details["category"] = None
    if raise_on_fail:
        raise ValueError(
            "Типология 'park' НЕ подтверждена: точка не входит ни в один полигон "
            "leisure=park или natural=wood (по данным OSM). Выберите другую точку."
        )
    return details


# Модуль предоставляет функции классификации и экспорта.
# Основные функции: classify_point(), export_classification_csv(), classify_and_export_csv()


def find_header_columns(ws, required_headers, search_rows=15):
    """Find header row and map header names to column numbers."""
    required_lower = {h.lower(): h for h in required_headers}

    for r in range(1, search_rows + 1):
        mapping = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str):
                key = v.strip().lower()
                if key in required_lower:
                    mapping[required_lower[key]] = c

        if all(h in mapping for h in required_headers):
            return r, mapping

    raise ValueError(f"Не найдена строка заголовков с полями {required_headers} в первых {search_rows} строках.")


def export_classification_csv(lat: float, lon: float, city: str, region: str, category: Optional[str], output_dir: str = "data/content/evaluation") -> str:
    """Export classification result to CSV file.

    Creates or appends to CSV file with pattern: {city}_class_{YYYY-MM-DD}.csv
    Returns the path to the created/updated CSV file.
    """
    os.makedirs(output_dir, exist_ok=True)

    today = datetime.now().strftime("%Y-%m-%d")
    filename = f"{city.lower()}_class_{today}.csv"
    filepath = os.path.join(output_dir, filename)

    headers = ["№", "city", "region", "name", "coordinate", "category"]
    data_row = [0, city, region, "object", f"{lat},{lon}", category or ""]

    file_exists = os.path.exists(filepath)

    with open(filepath, "a", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        if not file_exists:
            writer.writerow(headers)
        writer.writerow(data_row)

    return filepath


def classify_and_export_csv(lat: float, lon: float, city: str, region: str, R_CAT: int = 1500, output_dir: str = "data/content/evaluation") -> Dict[str, Any]:
    """Classify point and automatically export result to CSV.

    Combines classify_point() with CSV export. Returns classification result dict.
    Raises ValueError if classification fails and raise_on_fail=True.
    """
    result = classify_point(lat, lon, R_CAT=R_CAT, raise_on_fail=False)
    category = result.get("category")

    csv_path = export_classification_csv(lat, lon, city, region, category, output_dir)
    result["csv_path"] = csv_path

    return result